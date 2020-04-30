import requests
import json
import sys
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from app import Ui_MainWindow
from PyQt5.QtWidgets import *
from openpyxl.utils import get_column_letter


class Win(QMainWindow):
    def __init__(self):
        super(Win, self).__init__()

        self.ui = Ui_MainWindow()

        self.ui.setupUi(self)
        self.setWindowTitle('RChecker')
        self.ui.start.clicked.connect(self.start)
        self.ui.file.clicked.connect(self.getFileName)
        self.ui.login.setPlaceholderText('Логин')
        self.ui.psw.setPlaceholderText('Пароль')
        self.ui.label_ready.hide()
        self.ui.login.setText('Тома')
        self.filename = None


    def start(self):
        log = self.ui.login.text()
        psw = self.ui.psw.text()
        column = self.ui.column.text()
        # if log != 'Тома':    супер проверка лицензии
        #     sys.exit(app.exec())
        checking(log, psw, self.filename, column)

        self.show_ready_label()

    def show_ready_label(self):
        self.ui.label_ready.show()
        self.ui.label_column.hide()
        self.ui.column.hide()
        self.ui.label.hide()
        self.ui.psw.hide()
        self.ui.start.hide()
        self.ui.login.hide()
        self.ui.file.hide()
        self.ui.line.hide()

    def getFileName(self):
        self.filename, filetype = QFileDialog.getOpenFileName(self)

        self.check_file_name()

    def check_file_name(self):
        if self.filename != '':
            if self.filename[-4:] in ['xlsx', 'xlsm', 'xltx', 'xltm']:
                self.ui.file.setText('Файл выбран')
                self.ui.file.setStyleSheet("#file{\n"
                                           "border: 2px solid #3471FF;\n"
                                           "background-color:white;\n"
                                           "color:#3471FF}\n"
                                           "#file:hover{\n"
                                           "border: 2px solid #232A46;\n"
                                           "background-color:#232A46;\n"
                                           "color:white}\n"
                                           )

            else:

                self.ui.file.setStyleSheet("#file{\n"
                                           "border: 2px solid #DC546B;\n"
                                           "background-color:white;\n"
                                           "padding-botton: 5px;\n"
                                           "color:#DC546B}\n"

                                           "#file:hover{\n"
                                           "border: 2px solid #232A46;\n"
                                           "background-color:#232A46;\n"
                                           "color:white}\n"

                                           )

                self.ui.file.setText('Неверный\nформат')






def checking(log, psw, filename, column):
    site = WorkSite()
    excel = Work_excel()

    site.start_session(log, psw)
    excel_data = excel.load_data_from_excel(filename, column)
    site_data = site.get_data_from_site()
    ids = excel.comparing(excel_data, site_data)
    site.posting(ids)






class Work_excel:

    def load_data_from_excel(self, filename, column):
        self.filename = filename
        self.column = int(column)
        self.excel = load_workbook(filename)
        try:
            self.sheet = self.excel['Sheet1']
        except:
            self.sheet = self.excel['Лист1']
        rows = self.sheet.max_row
        data_from_sms = []
        for num_of_row in range(1, rows + 1):
            sms = self.sheet.cell(row=num_of_row, column=self.column).value.upper().split()
            status = self.sheet.cell(row=num_of_row, column=self.column + 1).value
            if sms[0] == 'ПЕРЕВОД' and status != 'ПРОВЕРЕНО':
                data_for_sms = [sms[3:6], sms[1]]
                data_for_sms[0] = ' '.join(sms[0])
                data_for_sms[1] = ''.join(sms[1])[:-1]
                data_for_sms.append(get_column_letter(self.column) + str(num_of_row))
                data_from_sms.append(data_for_sms)
        return data_from_sms

    def comparing(self, sms_data, site_data):
        ids = []
        yelFill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        for sms in sms_data:
            for site in site_data:
                if sms[0] == site[0] and int(sms[1]) == int(site[1]):
                    self.sheet[sms[2]].fill = yelFill

                    self.sheet[get_column_letter(self.column + 1) + sms[2][1:]].value = "ПРОВЕРЕНО"
                    sms_data.remove(sms)
                    ids.append(site[2])
        self.excel.save(self.filename)
        return ids


class WorkSite:
    def __init__(self):
        self.session = requests.Session()
        self.headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:68.0) Gecko/20100101 Firefox/68.0'}

    def start_session(self, log, pasw):

        url = 'https://rcm62.com/j_spring_security_check'
        data = {"j_username": log,
                "j_password": pasw,
                "_spring_security_remember_me": "on"}
        self.session.post(url, data=data, headers=self.headers)

    def get_data_from_site(self):
        num_of_page = 1
        datas_from_site = []
        while num_of_page:
            json_url = 'https://rcm62.com/admin/orders/all/?filter=%7B%22status%22:%22CHECK_PAY%22%7D&order=%7B%22id%22:%22DESC%22%7D&page='
            dictJson = self.session.get(json_url + str(num_of_page), headers=self.headers)
            dictJson = json.loads(dictJson.text)

            try:
                for g in range(len(dictJson)):
                    ide = (dictJson[g]['id'])
                    summ = (dictJson[g]['itogo'])
                    name = dictJson[g]['paymentMessages'][0]['name'].split()
                    name[2] = name[2][0] + '.'
                    name = ' '.join(name).upper()
                    data_from_site = [name, summ, ide]
                    datas_from_site.append(data_from_site)

            except:
                pass
            if len(dictJson) == 10:
                num_of_page += 1
            else:
                num_of_page = False
        return datas_from_site

    def posting(self, ids):
        headers2 = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:68.0) Gecko/20100101 Firefox/68.0',
                    'Accept': '*',
                    'Accept-Encoding': 'gzip, deflate, br',
                    'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
                    'Connection': 'keep-alive',
                    'Content-Length': '22',
                    'Content-Type': 'application/json;charset=utf-8'}

        for i in ids:
            self.session.post(
                'https://rcm62.com/admin/orders/update-status/?byIds=1&ids=%5B' + (str(i)) + '%5D&status=PAYED',
                data={'byIds': 1,
                      'ids': [i],
                      'status': 'PAYED'},
                headers=headers2)





app = QApplication([])
application = Win()
application.show()

sys.exit(app.exec())
